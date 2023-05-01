import io
from rest_framework.response import Response
from django.forms import ValidationError
from planning.models import Stock_log
from planning.serializer import Stock_log_Serializer   
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font,Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.http import FileResponse, HttpResponse



def add_stock_log(verified_data):
    Stock_log_instance=Stock_log_Serializer(data=verified_data)
    try:
        Stock_log_instance.is_valid(raise_exception=True)
        Stock_log_instance.save()
        return Stock_log_instance.data
    except ValidationError as e:
        return {'error': str(e)},

def get_excel(data, name):
    
    if len(data) == 0:
        return Response('There is no bom for this product id', status=400)
    
    
    wb = openpyxl.Workbook()
    ws = wb.active

    # set the sheet title
    ws.title = name
        # add title
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = name
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=len(data[0]))
    
    # to adjust each column size
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12

    ws.column_dimensions['E'].width = 75

    # add header row
    headers = list(data[0].keys())
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws['{}2'.format(col_letter)] = header

    # add data rows
    for row_num, row_data in enumerate(data, 3):
        for col_num, cell_value in enumerate(row_data.values(), 1):
            col_letter = get_column_letter(col_num)
            ws.cell(row=row_num, column=col_num, value=cell_value)

    # create table
    table = Table(displayName="Table1", ref="A2:{}{}".format(col_letter, row_num))
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    # modify the filename here
    filename = name + '.xlsx'
    content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    # write the file to a buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # create the HttpResponse object with the file and appropriate headers
    response = HttpResponse(output.read(), content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response